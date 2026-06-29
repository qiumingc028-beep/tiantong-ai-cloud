import os
import secrets
import hashlib
from datetime import datetime
from fastapi import FastAPI, Request, Response, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
import psycopg2
import redis

load_dotenv("/data/apps/tiantong-ai/.env")

app = FastAPI(title="天统AI云中台")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DATABASE_URL = os.getenv("DATABASE_URL")
REDIS_HOST = os.getenv("REDIS_HOST", "127.0.0.1")
REDIS_PORT = int(os.getenv("REDIS_PORT", "6379"))
REDIS_PASSWORD = os.getenv("REDIS_PASSWORD")
SESSION_TTL = 7 * 24 * 3600

def db_conn():
    return psycopg2.connect(DATABASE_URL)

def redis_conn():
    return redis.Redis(
        host=REDIS_HOST,
        port=REDIS_PORT,
        password=REDIS_PASSWORD,
        decode_responses=True
    )

def hash_password(password: str, salt: str = None) -> str:
    salt = salt or secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 120000)
    return f"pbkdf2_sha256${salt}${dk.hex()}"

def verify_password(password: str, password_hash: str) -> bool:
    try:
        method, salt, digest = password_hash.split("$")
        return hash_password(password, salt) == password_hash
    except Exception:
        return False

def ensure_tables():
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id SERIAL PRIMARY KEY,
        username VARCHAR(50) UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role VARCHAR(50) NOT NULL,
        display_name VARCHAR(100) NOT NULL,
        active BOOLEAN DEFAULT TRUE,
        created_at TIMESTAMPTZ DEFAULT NOW()
    );
    """)
    conn.commit()
    cur.close()
    conn.close()

ensure_tables()

def get_user_by_id(user_id: int):
    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT id, username, role, display_name, active FROM users WHERE id=%s",
        (user_id,)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return None
    return {
        "id": row[0],
        "username": row[1],
        "role": row[2],
        "display_name": row[3],
        "active": row[4],
    }

def require_user(request: Request):
    token = request.cookies.get("tiantong_session")
    if not token:
        raise HTTPException(status_code=401, detail="未登录")

    r = redis_conn()
    user_id = r.get(f"session:{token}")
    if not user_id:
        raise HTTPException(status_code=401, detail="登录已过期")

    user = get_user_by_id(int(user_id))
    if not user or not user["active"]:
        raise HTTPException(status_code=401, detail="账号无效")

    return user

def require_admin(request: Request):
    user = require_user(request)
    if user["role"] not in ["owner", "boss"]:
        raise HTTPException(status_code=403, detail="没有老板控制权限")
    return user

@app.get("/api/health")
def health():
    db_ok = False
    redis_ok = False

    try:
        conn = db_conn()
        cur = conn.cursor()
        cur.execute("SELECT 1;")
        cur.close()
        conn.close()
        db_ok = True
    except Exception as e:
        db_ok = str(e)

    try:
        r = redis_conn()
        redis_ok = r.ping()
    except Exception as e:
        redis_ok = str(e)

    return {
        "system": "天统AI云中台",
        "status": "running",
        "database": db_ok,
        "redis": redis_ok,
        "time": datetime.now().isoformat()
    }

@app.post("/api/login")
async def login(request: Request, response: Response):
    data = await request.json()
    username = data.get("username", "").strip()
    password = data.get("password", "")

    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT id, username, password_hash, role, display_name, active FROM users WHERE username=%s",
        (username,)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        raise HTTPException(status_code=401, detail="账号或密码错误")

    user_id, username, password_hash, role, display_name, active = row

    if not active or not verify_password(password, password_hash):
        raise HTTPException(status_code=401, detail="账号或密码错误")

    token = secrets.token_urlsafe(32)
    r = redis_conn()
    r.setex(f"session:{token}", SESSION_TTL, str(user_id))

    response.set_cookie(
        key="tiantong_session",
        value=token,
        httponly=True,
        samesite="lax",
        max_age=SESSION_TTL
    )

    return {
        "ok": True,
        "user": {
            "id": user_id,
            "username": username,
            "role": role,
            "display_name": display_name
        }
    }

@app.post("/api/logout")
def logout(request: Request, response: Response):
    token = request.cookies.get("tiantong_session")
    if token:
        redis_conn().delete(f"session:{token}")
    response.delete_cookie("tiantong_session")
    return {"ok": True}

@app.get("/api/me")
def me(request: Request):
    return require_user(request)

@app.get("/api/dashboard")
def dashboard(request: Request):
    user = require_user(request)

    return {
        "title": "老板驾驶舱",
        "user": user,
        "today_sales": 0,
        "today_profit": 0,
        "ad_spend": 0,
        "roi": 0,
        "orders": 0,
        "refunds": 0,
        "after_sales": 0,
        "stores": 60,
        "message": "员工登录系统已启动，下一步接入京东店铺数据。"
    }

@app.get("/api/users")
def list_users(request: Request):
    require_admin(request)

    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, username, role, display_name, active, created_at
        FROM users
        ORDER BY id ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [
        {
            "id": r[0],
            "username": r[1],
            "role": r[2],
            "display_name": r[3],
            "active": r[4],
            "created_at": r[5].isoformat() if r[5] else None
        }
        for r in rows
    ]

@app.post("/api/users")
async def create_user(request: Request):
    require_admin(request)

    data = await request.json()
    username = data.get("username", "").strip()
    display_name = data.get("display_name", "").strip()
    role = data.get("role", "").strip()

    allowed_roles = ["boss", "operator", "service", "designer", "ads"]
    if role not in allowed_roles:
        raise HTTPException(status_code=400, detail="角色不正确")

    if not username or not display_name:
        raise HTTPException(status_code=400, detail="用户名和姓名不能为空")

    password = secrets.token_urlsafe(10)

    conn = db_conn()
    cur = conn.cursor()

    cur.execute("SELECT id FROM users WHERE username=%s", (username,))
    if cur.fetchone():
        cur.close()
        conn.close()
        raise HTTPException(status_code=400, detail="用户名已存在")

    cur.execute(
        "INSERT INTO users(username, password_hash, role, display_name, active) VALUES(%s,%s,%s,%s,TRUE)",
        (username, hash_password(password), role, display_name)
    )

    conn.commit()
    cur.close()
    conn.close()

    return {
        "ok": True,
        "username": username,
        "display_name": display_name,
        "role": role,
        "password": password
    }

@app.post("/api/users/{user_id}/reset-password")
def reset_user_password(user_id: int, request: Request):
    require_admin(request)

    password = secrets.token_urlsafe(10)

    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET password_hash=%s WHERE id=%s RETURNING username",
        (hash_password(password), user_id)
    )
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="用户不存在")

    return {
        "ok": True,
        "username": row[0],
        "password": password
    }

@app.post("/api/users/{user_id}/toggle")
def toggle_user(user_id: int, request: Request):
    admin = require_admin(request)

    if admin["id"] == user_id:
        raise HTTPException(status_code=400, detail="不能停用自己")

    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET active = NOT active WHERE id=%s RETURNING username, active",
        (user_id,)
    )
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="用户不存在")

    return {
        "ok": True,
        "username": row[0],
        "active": row[1]
    }

# --- Store Management Module ---

def ensure_store_tables():
    conn = db_conn()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS stores (
        id SERIAL PRIMARY KEY,
        platform VARCHAR(50) NOT NULL DEFAULT 'jd',
        store_code VARCHAR(100) UNIQUE NOT NULL,
        store_name VARCHAR(200) NOT NULL,
        manager_user_id INTEGER REFERENCES users(id),
        active BOOLEAN DEFAULT TRUE,
        notes TEXT,
        created_at TIMESTAMPTZ DEFAULT NOW()
    );
    """)

    conn.commit()
    cur.close()
    conn.close()

ensure_store_tables()

@app.get("/api/store-users")
def store_users(request: Request):
    require_admin(request)

    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, username, role, display_name, active
        FROM users
        WHERE active = TRUE
        ORDER BY id ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [
        {
            "id": r[0],
            "username": r[1],
            "role": r[2],
            "display_name": r[3],
            "active": r[4],
        }
        for r in rows
    ]

@app.get("/api/stores")
def list_stores(request: Request):
    require_admin(request)

    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            s.id,
            s.platform,
            s.store_code,
            s.store_name,
            s.active,
            s.notes,
            s.created_at,
            u.id AS manager_id,
            u.username AS manager_username,
            u.display_name AS manager_name
        FROM stores s
        LEFT JOIN users u ON s.manager_user_id = u.id
        ORDER BY s.id ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [
        {
            "id": r[0],
            "platform": r[1],
            "store_code": r[2],
            "store_name": r[3],
            "active": r[4],
            "notes": r[5],
            "created_at": r[6].isoformat() if r[6] else None,
            "manager": {
                "id": r[7],
                "username": r[8],
                "display_name": r[9],
            } if r[7] else None
        }
        for r in rows
    ]

@app.post("/api/stores")
async def create_store(request: Request):
    require_admin(request)
    data = await request.json()

    platform = data.get("platform", "jd").strip()
    store_code = data.get("store_code", "").strip()
    store_name = data.get("store_name", "").strip()
    notes = data.get("notes", "").strip()

    if not store_code or not store_name:
        raise HTTPException(status_code=400, detail="店铺编号和店铺名称不能为空")

    conn = db_conn()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO stores(platform, store_code, store_name, notes, active)
            VALUES(%s,%s,%s,%s,TRUE)
            RETURNING id
        """, (platform, store_code, store_name, notes))
        store_id = cur.fetchone()[0]
        conn.commit()
    except Exception:
        conn.rollback()
        cur.close()
        conn.close()
        raise HTTPException(status_code=400, detail="店铺编号已存在或保存失败")

    cur.close()
    conn.close()

    return {"ok": True, "id": store_id}

@app.post("/api/stores/seed-jd")
def seed_jd_stores(request: Request):
    require_admin(request)

    conn = db_conn()
    cur = conn.cursor()

    created = 0
    for i in range(1, 61):
        code = f"JD{i:02d}"
        name = f"京东店铺{i:02d}"
        cur.execute("SELECT id FROM stores WHERE store_code=%s", (code,))
        exists = cur.fetchone()
        if not exists:
            cur.execute("""
                INSERT INTO stores(platform, store_code, store_name, active)
                VALUES('jd', %s, %s, TRUE)
            """, (code, name))
            created += 1

    conn.commit()
    cur.close()
    conn.close()

    return {"ok": True, "created": created, "message": f"已生成 {created} 个京东店铺占位"}

@app.post("/api/stores/{store_id}/assign")
async def assign_store(store_id: int, request: Request):
    require_admin(request)
    data = await request.json()

    manager_user_id = data.get("manager_user_id")

    conn = db_conn()
    cur = conn.cursor()

    if manager_user_id:
        cur.execute("SELECT id FROM users WHERE id=%s AND active=TRUE", (manager_user_id,))
        if not cur.fetchone():
            cur.close()
            conn.close()
            raise HTTPException(status_code=400, detail="负责人不存在或已停用")

        cur.execute("""
            UPDATE stores
            SET manager_user_id=%s
            WHERE id=%s
            RETURNING id
        """, (manager_user_id, store_id))
    else:
        cur.execute("""
            UPDATE stores
            SET manager_user_id=NULL
            WHERE id=%s
            RETURNING id
        """, (store_id,))

    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="店铺不存在")

    return {"ok": True}

@app.post("/api/stores/{store_id}/toggle")
def toggle_store(store_id: int, request: Request):
    require_admin(request)

    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE stores
        SET active = NOT active
        WHERE id=%s
        RETURNING store_name, active
    """, (store_id,))
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="店铺不存在")

    return {"ok": True, "store_name": row[0], "active": row[1]}

# --- JD Data Integration Module ---

def ensure_jd_integration_tables():
    conn = db_conn()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS jd_integrations (
        id SERIAL PRIMARY KEY,
        store_id INTEGER REFERENCES stores(id) ON DELETE CASCADE,
        source_type VARCHAR(50) NOT NULL,
        connection_mode VARCHAR(50) NOT NULL,
        merchant_id VARCHAR(100),
        app_key VARCHAR(200),
        status VARCHAR(50) DEFAULT 'pending',
        notes TEXT,
        active BOOLEAN DEFAULT TRUE,
        created_at TIMESTAMPTZ DEFAULT NOW(),
        updated_at TIMESTAMPTZ DEFAULT NOW()
    );
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS store_daily_metrics (
        id SERIAL PRIMARY KEY,
        store_id INTEGER REFERENCES stores(id) ON DELETE CASCADE,
        metric_date DATE NOT NULL,
        sales_amount NUMERIC(14,2) DEFAULT 0,
        profit_amount NUMERIC(14,2) DEFAULT 0,
        ad_spend NUMERIC(14,2) DEFAULT 0,
        roi NUMERIC(10,2) DEFAULT 0,
        orders_count INTEGER DEFAULT 0,
        visitors_count INTEGER DEFAULT 0,
        refunds_count INTEGER DEFAULT 0,
        after_sales_count INTEGER DEFAULT 0,
        source VARCHAR(50) DEFAULT 'manual',
        created_at TIMESTAMPTZ DEFAULT NOW(),
        UNIQUE(store_id, metric_date)
    );
    """)

    conn.commit()
    cur.close()
    conn.close()

ensure_jd_integration_tables()

@app.get("/api/jd/integrations")
def list_jd_integrations(request: Request):
    require_admin(request)

    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            i.id,
            i.store_id,
            s.store_code,
            s.store_name,
            i.source_type,
            i.connection_mode,
            i.merchant_id,
            i.app_key,
            i.status,
            i.notes,
            i.active,
            i.created_at,
            i.updated_at
        FROM jd_integrations i
        JOIN stores s ON i.store_id = s.id
        ORDER BY i.id ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [
        {
            "id": r[0],
            "store_id": r[1],
            "store_code": r[2],
            "store_name": r[3],
            "source_type": r[4],
            "connection_mode": r[5],
            "merchant_id": r[6],
            "app_key": r[7],
            "status": r[8],
            "notes": r[9],
            "active": r[10],
            "created_at": r[11].isoformat() if r[11] else None,
            "updated_at": r[12].isoformat() if r[12] else None
        }
        for r in rows
    ]

@app.post("/api/jd/integrations")
async def create_jd_integration(request: Request):
    require_admin(request)
    data = await request.json()

    store_id = data.get("store_id")
    source_type = data.get("source_type", "").strip()
    connection_mode = data.get("connection_mode", "").strip()
    merchant_id = data.get("merchant_id", "").strip()
    app_key = data.get("app_key", "").strip()
    notes = data.get("notes", "").strip()

    if not store_id or not source_type or not connection_mode:
        raise HTTPException(status_code=400, detail="店铺、数据来源、接入方式不能为空")

    allowed_sources = ["jd_sz", "jd_jzt", "jd_open", "manual_import", "browser_auto"]
    allowed_modes = ["official_api", "browser_auto", "excel_import", "pending"]

    if source_type not in allowed_sources:
        raise HTTPException(status_code=400, detail="数据来源不正确")

    if connection_mode not in allowed_modes:
        raise HTTPException(status_code=400, detail="接入方式不正确")

    conn = db_conn()
    cur = conn.cursor()

    cur.execute("SELECT id FROM stores WHERE id=%s", (store_id,))
    if not cur.fetchone():
        cur.close()
        conn.close()
        raise HTTPException(status_code=404, detail="店铺不存在")

    cur.execute("""
        INSERT INTO jd_integrations(
            store_id, source_type, connection_mode, merchant_id, app_key, notes, status, active
        )
        VALUES(%s,%s,%s,%s,%s,%s,'pending',TRUE)
        RETURNING id
    """, (store_id, source_type, connection_mode, merchant_id, app_key, notes))

    integration_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    conn.close()

    return {"ok": True, "id": integration_id}

@app.post("/api/jd/integrations/{integration_id}/status")
async def update_jd_integration_status(integration_id: int, request: Request):
    require_admin(request)
    data = await request.json()
    status = data.get("status", "").strip()

    allowed = ["pending", "authorized", "error", "disabled"]
    if status not in allowed:
        raise HTTPException(status_code=400, detail="状态不正确")

    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE jd_integrations
        SET status=%s, updated_at=NOW()
        WHERE id=%s
        RETURNING id
    """, (status, integration_id))
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="接入记录不存在")

    return {"ok": True}

@app.post("/api/jd/integrations/{integration_id}/toggle")
def toggle_jd_integration(integration_id: int, request: Request):
    require_admin(request)

    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE jd_integrations
        SET active = NOT active, updated_at=NOW()
        WHERE id=%s
        RETURNING id, active
    """, (integration_id,))
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="接入记录不存在")

    return {"ok": True, "active": row[1]}

@app.get("/api/jd/metrics/summary")
def jd_metrics_summary(request: Request):
    require_user(request)

    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            COALESCE(SUM(sales_amount),0),
            COALESCE(SUM(profit_amount),0),
            COALESCE(SUM(ad_spend),0),
            COALESCE(SUM(orders_count),0),
            COALESCE(SUM(visitors_count),0),
            COALESCE(SUM(refunds_count),0),
            COALESCE(SUM(after_sales_count),0)
        FROM store_daily_metrics
        WHERE metric_date = CURRENT_DATE
    """)
    r = cur.fetchone()
    cur.close()
    conn.close()

    sales = float(r[0] or 0)
    ad = float(r[2] or 0)
    roi = round(sales / ad, 2) if ad > 0 else 0

    return {
        "today_sales": sales,
        "today_profit": float(r[1] or 0),
        "ad_spend": ad,
        "roi": roi,
        "orders": int(r[3] or 0),
        "visitors": int(r[4] or 0),
        "refunds": int(r[5] or 0),
        "after_sales": int(r[6] or 0)
    }

# --- Manual Metrics Module ---

@app.post("/api/metrics/manual")
async def save_manual_metrics(request: Request):
    require_admin(request)
    data = await request.json()

    store_id = data.get("store_id")
    metric_date = data.get("metric_date")
    sales_amount = data.get("sales_amount", 0) or 0
    profit_amount = data.get("profit_amount", 0) or 0
    ad_spend = data.get("ad_spend", 0) or 0
    roi = data.get("roi", 0) or 0
    orders_count = data.get("orders_count", 0) or 0
    visitors_count = data.get("visitors_count", 0) or 0
    refunds_count = data.get("refunds_count", 0) or 0
    after_sales_count = data.get("after_sales_count", 0) or 0

    if not store_id:
        raise HTTPException(status_code=400, detail="请选择店铺")

    if not metric_date:
        metric_date = datetime.now().date().isoformat()

    conn = db_conn()
    cur = conn.cursor()

    cur.execute("SELECT id FROM stores WHERE id=%s", (store_id,))
    if not cur.fetchone():
        cur.close()
        conn.close()
        raise HTTPException(status_code=404, detail="店铺不存在")

    cur.execute("""
        INSERT INTO store_daily_metrics(
            store_id,
            metric_date,
            sales_amount,
            profit_amount,
            ad_spend,
            roi,
            orders_count,
            visitors_count,
            refunds_count,
            after_sales_count,
            source
        )
        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'manual')
        ON CONFLICT(store_id, metric_date)
        DO UPDATE SET
            sales_amount=EXCLUDED.sales_amount,
            profit_amount=EXCLUDED.profit_amount,
            ad_spend=EXCLUDED.ad_spend,
            roi=EXCLUDED.roi,
            orders_count=EXCLUDED.orders_count,
            visitors_count=EXCLUDED.visitors_count,
            refunds_count=EXCLUDED.refunds_count,
            after_sales_count=EXCLUDED.after_sales_count,
            source='manual';
    """, (
        store_id,
        metric_date,
        sales_amount,
        profit_amount,
        ad_spend,
        roi,
        orders_count,
        visitors_count,
        refunds_count,
        after_sales_count
    ))

    conn.commit()
    cur.close()
    conn.close()

    return {"ok": True, "message": "数据已保存"}

@app.get("/api/metrics/today")
def metrics_today(request: Request):
    require_user(request)

    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            s.id,
            s.store_code,
            s.store_name,
            COALESCE(m.sales_amount, 0),
            COALESCE(m.profit_amount, 0),
            COALESCE(m.ad_spend, 0),
            COALESCE(m.roi, 0),
            COALESCE(m.orders_count, 0),
            COALESCE(m.visitors_count, 0),
            COALESCE(m.refunds_count, 0),
            COALESCE(m.after_sales_count, 0)
        FROM stores s
        LEFT JOIN store_daily_metrics m
            ON s.id = m.store_id
            AND m.metric_date = CURRENT_DATE
        WHERE s.active = TRUE
        ORDER BY s.id ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [
        {
            "store_id": r[0],
            "store_code": r[1],
            "store_name": r[2],
            "sales_amount": float(r[3] or 0),
            "profit_amount": float(r[4] or 0),
            "ad_spend": float(r[5] or 0),
            "roi": float(r[6] or 0),
            "orders_count": int(r[7] or 0),
            "visitors_count": int(r[8] or 0),
            "refunds_count": int(r[9] or 0),
            "after_sales_count": int(r[10] or 0)
        }
        for r in rows
    ]

# --- Batch Metrics Import Module ---
from fastapi import UploadFile, File
from io import BytesIO, StringIO
import csv
from datetime import date
import openpyxl

def safe_number(v):
    if v is None or v == "":
        return 0
    try:
        return float(v)
    except Exception:
        return 0

def safe_int(v):
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except Exception:
        return 0

def normalize_date(v):
    """把 Excel / CSV 日期统一转换成 YYYY-MM-DD。
    支持：Excel日期对象、Excel序列号如46202、文本日期。
    """
    if v is None or v == "":
        return datetime.now().date().isoformat()

    if hasattr(v, "date"):
        return v.date().isoformat()

    # Excel 日期序列号，例如 46202
    try:
        if isinstance(v, (int, float)) or str(v).strip().isdigit():
            from datetime import datetime as _dt, timedelta as _timedelta
            n = int(float(v))
            if 20000 <= n <= 60000:
                return (_dt(1899, 12, 30) + _timedelta(days=n)).date().isoformat()
    except Exception:
        pass

    text = str(v).strip()
    text = text.replace("/", "-").replace(".", "-")

    try:
        from datetime import datetime as _dt

        # 支持 20260629
        if len(text) == 8 and text.isdigit():
            return _dt.strptime(text, "%Y%m%d").date().isoformat()

        # 支持 2026-6-29 / 2026-06-29
        return _dt.fromisoformat(text).date().isoformat()
    except Exception:
        return datetime.now().date().isoformat()


def get_value(row, *names):
    for n in names:
        if n in row:
            return row.get(n)
    return None

@app.post("/api/metrics/import")
async def import_metrics_file(request: Request, file: UploadFile = File(...)):
    require_admin(request)

    content = await file.read()
    filename = (file.filename or "").lower()
    rows = []

    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(r):
                continue
            row = {}
            for idx, h in enumerate(headers):
                row[h] = r[idx] if idx < len(r) else None
            rows.append(row)

    elif filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        rows = list(reader)
    else:
        raise HTTPException(status_code=400, detail="只支持 .xlsx 或 .csv 文件")

    conn = db_conn()
    cur = conn.cursor()

    imported = 0
    errors = []

    for index, row in enumerate(rows, start=2):
        store_code = str(get_value(row, "店铺编号", "store_code", "编号") or "").strip()
        if not store_code:
            errors.append(f"第{index}行：缺少店铺编号")
            continue

        cur.execute("SELECT id FROM stores WHERE store_code=%s", (store_code,))
        store = cur.fetchone()
        if not store:
            errors.append(f"第{index}行：找不到店铺编号 {store_code}")
            continue

        store_id = store[0]
        metric_date = normalize_date(get_value(row, "日期", "metric_date", "date"))

        sales_amount = safe_number(get_value(row, "今日成交", "成交", "sales_amount"))
        profit_amount = safe_number(get_value(row, "今日利润", "利润", "profit_amount"))
        ad_spend = safe_number(get_value(row, "广告花费", "广告费", "ad_spend"))
        roi = safe_number(get_value(row, "ROI", "roi"))
        orders_count = safe_int(get_value(row, "订单数", "订单", "orders_count"))
        visitors_count = safe_int(get_value(row, "访客数", "访客", "visitors_count"))
        refunds_count = safe_int(get_value(row, "退款数", "退款", "refunds_count"))
        after_sales_count = safe_int(get_value(row, "售后数", "售后", "after_sales_count"))

        cur.execute("""
            INSERT INTO store_daily_metrics(
                store_id,
                metric_date,
                sales_amount,
                profit_amount,
                ad_spend,
                roi,
                orders_count,
                visitors_count,
                refunds_count,
                after_sales_count,
                source
            )
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'excel')
            ON CONFLICT(store_id, metric_date)
            DO UPDATE SET
                sales_amount=EXCLUDED.sales_amount,
                profit_amount=EXCLUDED.profit_amount,
                ad_spend=EXCLUDED.ad_spend,
                roi=EXCLUDED.roi,
                orders_count=EXCLUDED.orders_count,
                visitors_count=EXCLUDED.visitors_count,
                refunds_count=EXCLUDED.refunds_count,
                after_sales_count=EXCLUDED.after_sales_count,
                source='excel';
        """, (
            store_id,
            metric_date,
            sales_amount,
            profit_amount,
            ad_spend,
            roi,
            orders_count,
            visitors_count,
            refunds_count,
            after_sales_count
        ))

        imported += 1

    conn.commit()
    cur.close()
    conn.close()

    return {
        "ok": True,
        "imported": imported,
        "errors": errors
    }
