import csv
import hashlib
import json
import math
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from threading import Lock

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from sqlalchemy.orm import Session

from ..auth import current_user, require_permission_user
from ..database import get_db, get_redis
from ..models import AiTask, EmployeeLog, JdDailyMetric, MetricDaily, Store


router = APIRouter()
MAX_IMPORT_FILE_SIZE = 512 * 1024
IMPORT_LOCK = Lock()


@router.get("/api/jd/metrics/summary")
def jd_metrics_summary(request: Request, db: Session = Depends(get_db)):
    current_user(request, db)
    return metrics_summary(db)


@router.get("/api/owner/dashboard")
def owner_dashboard(
    request: Request,
    store_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    db: Session = Depends(get_db),
):
    user = require_permission_user(request, db, "menu.dashboard")
    data = metrics_summary(db, store_id, date_from, date_to)
    data["user"] = {"id": user.id, "username": user.username, "display_name": user.display_name, "role": user.role}
    data["title"] = "老板驾驶舱"
    return data


@router.post("/api/metrics/manual")
async def save_manual_metrics(request: Request, db: Session = Depends(get_db)):
    user = require_permission_user(request, db, "data.metrics.write")
    data = await request.json()
    store_id = data.get("store_id")
    if not store_id:
        raise HTTPException(status_code=400, detail="请选择店铺")
    if not db.get(Store, store_id):
        raise HTTPException(status_code=404, detail="店铺不存在")
    metric_date = parse_date(data.get("metric_date")) or date.today()
    upsert_metric(db, store_id, metric_date, data, "manual", user.id)
    upsert_jd_daily_from_manual(db, store_id, metric_date, data)
    db.commit()
    return {"ok": True, "message": "数据已保存"}


@router.get("/api/metrics/today")
def metrics_today(request: Request, db: Session = Depends(get_db)):
    current_user(request, db)
    jd_metrics = {m.store_id: m for m in db.query(JdDailyMetric).filter(JdDailyMetric.metric_date == date.today()).all()}
    legacy_metrics = {m.store_id: m for m in db.query(MetricDaily).filter(MetricDaily.metric_date == date.today()).all()}
    stores = db.query(Store).filter(Store.active.is_(True)).order_by(Store.id.asc()).all()
    return [metric_row(store, jd_metrics.get(store.id), legacy_metrics.get(store.id)) for store in stores]


@router.get("/api/business-center/metrics")
def business_center_metrics(
    request: Request,
    store_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    db: Session = Depends(get_db),
):
    require_permission_user(request, db, "menu.jd_data")
    rows = business_metric_rows(db, store_id, date_from, date_to)
    return {"rows": rows, "total": len(rows), "summary": summarize_business_rows(rows)}


@router.get("/api/metrics/import-records")
def import_records(request: Request, db: Session = Depends(get_db)):
    user = require_permission_user(request, db, "data.metrics.write")
    logs = (
        db.query(EmployeeLog)
        .filter(EmployeeLog.user_id == user.id, EmployeeLog.action.like("metrics_import:%"))
        .order_by(EmployeeLog.id.desc())
        .limit(50)
        .all()
    )
    return {
        "records": [
            {**json.loads(log.detail or "{}"), "created_at": log.created_at.isoformat() if log.created_at else None}
            for log in logs
        ]
    }


@router.post("/api/metrics/import")
async def import_metrics_file(
    request: Request,
    file: UploadFile = File(...),
    store_id: int | None = Form(None),
    db: Session = Depends(get_db),
):
    user = require_permission_user(request, db, "data.metrics.write")
    content = await file.read()
    if len(content) > MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status_code=413, detail="导入文件过大")
    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx"):
        rows = read_xlsx_rows(content)
    elif filename.endswith(".csv"):
        rows = list(csv.DictReader(StringIO(content.decode("utf-8-sig"))))
    else:
        raise HTTPException(status_code=400, detail="只支持 .xlsx 或 .csv 文件")

    validate_import_schema(rows)
    # ponytail: one process lock plus DB row locks; add a unique idempotency key only if import throughput grows.
    with IMPORT_LOCK:
        return persist_metric_import(db, user, content, rows, store_id)


def persist_metric_import(db: Session, user, content: bytes, rows: list[dict], store_id: int | None):
    selected_store = None
    if store_id is not None:
        selected_store = db.query(Store).filter(Store.id == store_id).with_for_update().one_or_none()
    else:
        store_codes = sorted(
            {
                str(get_value(row, "店铺编号", "store_code", "编号") or "").strip()
                for row in rows
                if get_value(row, "店铺编号", "store_code", "编号")
            }
        )
        if store_codes:
            db.query(Store).filter(Store.store_code.in_(store_codes)).order_by(Store.id).with_for_update().all()

    if store_id is not None and (not selected_store or not selected_store.active):
        raise HTTPException(status_code=404, detail="店铺不存在或已停用")
    import_key = hashlib.sha256(f"{user.id}:{store_id or 0}:".encode() + content).hexdigest()
    action = f"metrics_import:{import_key}"
    previous = db.query(EmployeeLog).filter(EmployeeLog.user_id == user.id, EmployeeLog.action == action).one_or_none()
    if previous:
        result = json.loads(previous.detail or "{}")
        result["duplicate"] = True
        return result

    imported = 0
    errors = []
    for index, row in enumerate(rows, start=2):
        store_code = str(get_value(row, "店铺编号", "store_code", "编号") or "").strip()
        if selected_store and store_code and store_code != selected_store.store_code:
            errors.append({"row": index, "reason": "店铺与当前选择不一致"})
            continue
        store = selected_store or db.query(Store).filter(Store.store_code == store_code).one_or_none()
        if not store:
            errors.append({"row": index, "reason": "缺少或找不到店铺编号"})
            continue
        try:
            validate_required_import_values(row)
            data = parse_import_values(row)
        except ValueError as exc:
            errors.append({"row": index, "reason": str(exc)})
            continue
        raw_date = get_value(row, "日期", "metric_date", "date")
        metric_date = parse_date(raw_date)
        if raw_date and not metric_date:
            errors.append({"row": index, "reason": "日期格式错误"})
            continue
        upsert_metric(db, store.id, metric_date, data, "excel", user.id)
        upsert_jd_daily_from_manual(db, store.id, metric_date, data, "excel")
        imported += 1
    failed = len(errors)
    status = "success" if imported and not failed else ("partial_success" if imported else "failed")
    result = {
        "ok": imported > 0,
        "status": status,
        "total_rows": len(rows),
        "success_rows": imported,
        "failed_rows": failed,
        "imported": imported,
        "errors": errors,
        "duplicate": False,
    }
    db.add(EmployeeLog(user_id=user.id, action=action, detail=json.dumps(result, ensure_ascii=False)))
    db.commit()
    return result


def business_metric_rows(
    db: Session,
    store_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
):
    if date_from and date_to and date_from > date_to:
        raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")
    if store_id and not db.get(Store, store_id):
        raise HTTPException(status_code=404, detail="店铺不存在")
    if not date_from and not date_to:
        date_from = date_to = date.today()
    query = db.query(JdDailyMetric, Store).join(Store, Store.id == JdDailyMetric.store_id)
    if store_id:
        query = query.filter(JdDailyMetric.store_id == store_id)
    if date_from:
        query = query.filter(JdDailyMetric.metric_date >= date_from)
    if date_to:
        query = query.filter(JdDailyMetric.metric_date <= date_to)
    return [
        {
            "metric_date": metric.metric_date.isoformat(),
            "store_id": store.id,
            "store_name": store.store_name,
            "sales_amount": float(metric.gmv or 0),
            "profit_amount": float(metric.profit_amount or 0),
            "orders_count": int(metric.paid_orders_count or 0),
            "ad_spend": float(metric.ad_spend or 0),
            "visitors_count": int(metric.visitors_count or 0),
            "favorites_count": int(metric.favorites_count or 0),
            "cart_add_count": int(metric.cart_add_count or 0),
            "conversion_rate": float(metric.conversion_rate or 0),
            "refunds_count": int(metric.refunds_count or 0),
            "after_sales_count": int(metric.after_sales_count or 0),
        }
        for metric, store in query.order_by(JdDailyMetric.metric_date.desc(), Store.id.asc()).all()
    ]


def summarize_business_rows(rows):
    return {
        "sales_amount": round(sum(row["sales_amount"] for row in rows), 2),
        "orders_count": sum(row["orders_count"] for row in rows),
        "ad_spend": round(sum(row["ad_spend"] for row in rows), 2),
        "visitors_count": sum(row["visitors_count"] for row in rows),
        "favorites_count": sum(row["favorites_count"] for row in rows),
        "cart_add_count": sum(row["cart_add_count"] for row in rows),
    }


def metrics_summary(db: Session, store_id: int | None = None, date_from: date | None = None, date_to: date | None = None):
    rows = business_metric_rows(db, store_id, date_from, date_to)
    summary = summarize_business_rows(rows)
    gmv = summary["sales_amount"]
    ad = summary["ad_spend"]
    ai_tasks = db.query(AiTask).order_by(AiTask.id.asc()).all()
    return {
        "today_sales": gmv,
        "today_gmv": gmv,
        "today_profit": round(sum(row["profit_amount"] for row in rows), 2),
        "ad_spend": ad,
        "roi": round(gmv / ad, 2) if ad > 0 else 0,
        "orders": summary["orders_count"],
        "visitors": summary["visitors_count"],
        "refunds": sum(row["refunds_count"] for row in rows),
        "after_sales": sum(row["after_sales_count"] for row in rows),
        "stores": db.query(Store).filter(Store.platform == "jd", Store.active.is_(True)).count(),
        "ai_employees_online": count_online_employees(),
        "ai_task_status": [{"name": t.ai_employee_name, "status": t.status} for t in ai_tasks],
    }


def upsert_metric(db: Session, store_id: int, metric_date: date, data: dict, source: str, user_id: int):
    metric = db.query(MetricDaily).filter(MetricDaily.store_id == store_id, MetricDaily.metric_date == metric_date).one_or_none()
    if not metric:
        metric = MetricDaily(store_id=store_id, metric_date=metric_date)
        db.add(metric)
    metric.sales_amount = safe_number(data.get("sales_amount"))
    metric.profit_amount = safe_number(data.get("profit_amount"))
    metric.ad_spend = safe_number(data.get("ad_spend"))
    metric.roi = safe_number(data.get("roi"))
    metric.orders_count = safe_int(data.get("orders_count"))
    metric.visitors_count = safe_int(data.get("visitors_count"))
    metric.refunds_count = safe_int(data.get("refunds_count"))
    metric.after_sales_count = safe_int(data.get("after_sales_count"))
    metric.source = source
    metric.created_by = user_id


def upsert_jd_daily_from_manual(db: Session, store_id: int, metric_date: date, data: dict, source: str = "manual"):
    metric = db.query(JdDailyMetric).filter(JdDailyMetric.store_id == store_id, JdDailyMetric.metric_date == metric_date).one_or_none()
    if not metric:
        metric = JdDailyMetric(store_id=store_id, metric_date=metric_date)
        db.add(metric)
    metric.gmv = safe_number(data.get("sales_amount"))
    metric.profit_amount = safe_number(data.get("profit_amount"))
    metric.ad_spend = safe_number(data.get("ad_spend"))
    metric.roi = safe_number(data.get("roi"))
    metric.paid_orders_count = safe_int(data.get("orders_count"))
    metric.visitors_count = safe_int(data.get("visitors_count"))
    metric.refunds_count = safe_int(data.get("refunds_count"))
    metric.after_sales_count = safe_int(data.get("after_sales_count"))
    metric.favorites_count = safe_int(data.get("favorites_count"))
    metric.cart_add_count = safe_int(data.get("cart_add_count"))
    metric.conversion_rate = safe_number(data.get("conversion_rate"))
    metric.source = source


def metric_row(store: Store, jd_metric: JdDailyMetric | None, legacy_metric: MetricDaily | None):
    return {
        "store_id": store.id,
        "store_code": store.store_code,
        "store_name": store.store_name,
        "sales_amount": float(jd_metric.gmv if jd_metric else (legacy_metric.sales_amount if legacy_metric else 0)),
        "profit_amount": float(jd_metric.profit_amount if jd_metric else (legacy_metric.profit_amount if legacy_metric else 0)),
        "ad_spend": float(jd_metric.ad_spend if jd_metric else (legacy_metric.ad_spend if legacy_metric else 0)),
        "roi": float(jd_metric.roi if jd_metric else (legacy_metric.roi if legacy_metric else 0)),
        "orders_count": int(jd_metric.paid_orders_count if jd_metric else (legacy_metric.orders_count if legacy_metric else 0)),
        "visitors_count": int(jd_metric.visitors_count if jd_metric else (legacy_metric.visitors_count if legacy_metric else 0)),
        "refunds_count": int(jd_metric.refunds_count if jd_metric else (legacy_metric.refunds_count if legacy_metric else 0)),
        "after_sales_count": int(jd_metric.after_sales_count if jd_metric else (legacy_metric.after_sales_count if legacy_metric else 0)),
    }


def read_xlsx_rows(content):
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    return [{h: values[idx] if idx < len(values) else None for idx, h in enumerate(headers)} for values in ws.iter_rows(min_row=2, values_only=True) if any(values)]


def validate_import_schema(rows):
    if not rows:
        return
    headers = set(rows[0])
    required = (
        ("日期", ("日期", "metric_date", "date")),
        ("销售额", ("今日成交", "成交", "sales_amount")),
        ("订单量", ("订单数", "订单", "orders_count")),
        ("广告消耗", ("广告花费", "广告费", "ad_spend")),
    )
    missing = [label for label, aliases in required if not headers.intersection(aliases)]
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必要字段：{'、'.join(missing)}")


def validate_required_import_values(row):
    required = (
        ("日期", ("日期", "metric_date", "date")),
        ("销售额", ("今日成交", "成交", "sales_amount")),
        ("订单量", ("订单数", "订单", "orders_count")),
        ("广告消耗", ("广告花费", "广告费", "ad_spend")),
    )
    for label, aliases in required:
        if get_value(row, *aliases) in (None, ""):
            raise ValueError(f"{label}不能为空")


def parse_import_values(row):
    fields = (
        ("sales_amount", "销售额", ("今日成交", "成交", "sales_amount"), False),
        ("profit_amount", "利润", ("今日利润", "利润", "profit_amount"), False),
        ("ad_spend", "广告消耗", ("广告花费", "广告费", "ad_spend"), False),
        ("roi", "ROI", ("ROI", "roi"), False),
        ("orders_count", "订单量", ("订单数", "订单", "orders_count"), True),
        ("visitors_count", "访客数", ("访客数", "访客", "visitors_count"), True),
        ("refunds_count", "退款数", ("退款数", "退款", "refunds_count"), True),
        ("after_sales_count", "售后数", ("售后数", "售后", "after_sales_count"), True),
        ("favorites_count", "收藏数", ("收藏", "favorites_count"), True),
        ("cart_add_count", "加购数", ("加购", "cart_add_count"), True),
        ("conversion_rate", "转化率", ("转化率", "conversion_rate"), False),
    )
    parsed = {}
    for key, label, aliases, integer in fields:
        value = get_value(row, *aliases)
        if value in (None, ""):
            parsed[key] = 0
            continue
        try:
            number = float(value)
        except (TypeError, ValueError):
            raise ValueError(f"{label}格式错误") from None
        if not math.isfinite(number) or number < 0 or (integer and not number.is_integer()):
            raise ValueError(f"{label}格式错误")
        parsed[key] = int(number) if integer else number
    return parsed


def safe_number(v):
    try:
        return float(v or 0)
    except Exception:
        return 0


def safe_int(v):
    try:
        return int(float(v or 0))
    except Exception:
        return 0


def parse_date(v):
    if not v:
        return None
    if hasattr(v, "date"):
        return v.date()
    text = str(v).strip().replace("/", "-").replace(".", "-")
    try:
        if len(text) == 8 and text.isdigit():
            return datetime.strptime(text, "%Y%m%d").date()
        if text.isdigit():
            n = int(float(text))
            if 20000 <= n <= 60000:
                return (datetime(1899, 12, 30) + timedelta(days=n)).date()
        return datetime.fromisoformat(text).date()
    except Exception:
        return None


def get_value(row, *names):
    for name in names:
        if name in row:
            return row.get(name)
    return None


def count_online_employees():
    try:
        return sum(1 for _ in get_redis().scan_iter("session:*"))
    except Exception:
        return 0
