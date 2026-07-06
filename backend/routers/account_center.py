from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session, joinedload

from ..auth import hash_password, require_permission_user
from ..database import get_db
from ..models import (
    AccountTemplate,
    Brand,
    EmployeeStoreAssignment,
    Store,
    StoreAccountNote,
    StoreGroup,
)


router = APIRouter()

ACCOUNT_STATUSES = {"正常", "异常", "待登录", "Cookie过期", "暂停采集"}
COOKIE_STATUSES = {"正常", "异常", "待登录", "Cookie过期", "暂停采集", "有效", "失效", "未配置", "待更新"}
ACCOUNT_HEADERS = [
    "店铺编号",
    "店铺名称",
    "品牌",
    "负责人",
    "手机号",
    "登录账号",
    "Cookie状态",
    "登录状态",
    "最后登录时间",
    "备注",
    "标签",
]
ALLOWED_EXPORT_KEYS = (
    "store_code",
    "store_name",
    "brand",
    "owner",
    "phone",
    "login_account",
    "cookie_status",
    "login_status",
    "last_login_at",
    "notes",
    "tags",
)
SENSITIVE_FIELD_KEYWORDS = ("password", "encrypted_password", "cookie", "token", "api_key", "secret")
SENSITIVE_EXCEL_HEADERS = SENSITIVE_FIELD_KEYWORDS + ("密码", "口令", "密钥", "令牌")


@router.get("/api/accounts")
def list_accounts(
    request: Request,
    brand: str | None = None,
    owner: str | None = None,
    status: str | None = None,
    db: Session = Depends(get_db),
):
    require_account_user(request, db)
    query = (
        db.query(StoreAccountNote)
        .options(
            joinedload(StoreAccountNote.store),
            joinedload(StoreAccountNote.brand),
            joinedload(StoreAccountNote.group),
        )
        .filter(StoreAccountNote.active.is_(True))
        .join(Store, Store.id == StoreAccountNote.store_id)
        .outerjoin(Brand, Brand.id == StoreAccountNote.brand_id)
        .order_by(Store.store_code.asc())
    )
    if brand:
        query = query.filter(Brand.brand_name == brand)
    if status:
        query = query.filter(StoreAccountNote.account_status == status)

    notes = query.all()
    store_ids = [note.store_id for note in notes]
    assignments = {}
    if store_ids:
        assignments = {
            row.store_id: row
            for row in db.query(EmployeeStoreAssignment)
            .filter(EmployeeStoreAssignment.store_id.in_(store_ids))
            .all()
        }
    rows = [account_to_dict(note, assignments.get(note.store_id)) for note in notes]
    if owner:
        rows = [row for row in rows if row["owner"] == owner]
    return rows


@router.post("/api/accounts")
async def create_account(request: Request, db: Session = Depends(get_db)):
    require_account_user(request, db)
    data = await request.json()
    reject_forbidden_sensitive_payload(data, allow_plain_password=True)
    store_code = clean(data.get("store_code") or data.get("店铺编号"))
    store_name = clean(data.get("store_name") or data.get("店铺名称"))
    if not store_code or not store_name:
        raise HTTPException(status_code=400, detail="店铺编号和店铺名称不能为空")
    if db.query(Store).filter(Store.store_code == store_code).first():
        raise HTTPException(status_code=400, detail="店铺编号已存在")

    store = Store(platform="jd", store_code=store_code, store_name=store_name, active=True)
    db.add(store)
    db.flush()
    upsert_account_meta(db, store, data)
    db.commit()
    return {"ok": True, "id": store.id}


@router.put("/api/accounts/{account_id}")
async def update_account(account_id: int, request: Request, db: Session = Depends(get_db)):
    require_account_user(request, db)
    data = await request.json()
    reject_forbidden_sensitive_payload(data, allow_plain_password=True)
    store = db.get(Store, account_id)
    if not store:
        raise HTTPException(status_code=404, detail="账号资料不存在")

    store_code = clean(data.get("store_code") or data.get("店铺编号"))
    store_name = clean(data.get("store_name") or data.get("店铺名称"))
    if store_code and store_code != store.store_code:
        exists = db.query(Store).filter(Store.store_code == store_code, Store.id != store.id).first()
        if exists:
            raise HTTPException(status_code=400, detail="店铺编号已存在")
        store.store_code = store_code
    if store_name:
        store.store_name = store_name
    upsert_account_meta(db, store, data)
    db.commit()
    return {"ok": True, "id": store.id}


@router.delete("/api/accounts/{account_id}")
def delete_account(account_id: int, request: Request, db: Session = Depends(get_db)):
    require_account_user(request, db)
    note = db.query(StoreAccountNote).filter(StoreAccountNote.store_id == account_id).first()
    if not note:
        raise HTTPException(status_code=404, detail="账号资料不存在")
    note.active = False
    note.account_status = "暂停采集"
    db.commit()
    return {"ok": True, "id": account_id}


@router.post("/api/accounts/import")
async def import_accounts(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    require_account_user(request, db)
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="导入文件为空")
    try:
        workbook = load_workbook(BytesIO(content))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="请上传 xlsx 格式的 Excel 文件") from exc

    sheet = workbook.active
    headers = [clean(cell.value) for cell in sheet[1]]
    reject_sensitive_headers(headers)
    if "店铺编号" not in headers or "店铺名称" not in headers:
        raise HTTPException(status_code=400, detail="Excel 必须包含店铺编号和店铺名称")

    imported = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        store_code = clean(data.get("店铺编号"))
        store_name = clean(data.get("店铺名称"))
        if not store_code or not store_name:
            continue
        store = db.query(Store).filter(Store.store_code == store_code).first()
        if not store:
            store = Store(platform="jd", store_code=store_code, store_name=store_name, active=True)
            db.add(store)
            db.flush()
        else:
            store.store_name = store_name
        upsert_account_meta(db, store, data)
        imported += 1
    db.commit()
    return {"ok": True, "imported": imported}


@router.get("/api/accounts/export")
def export_accounts(request: Request, db: Session = Depends(get_db)):
    require_account_user(request, db)
    rows = list_accounts(request, db=db)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "账号资料"
    sheet.append(filter_safe_export_headers(ACCOUNT_HEADERS))
    for row in rows:
        safe_row = sanitize_export_row(row)
        sheet.append([safe_row[key] for key in ALLOWED_EXPORT_KEYS])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="jd_60_store_accounts.xlsx"'},
    )


@router.get("/api/templates")
def list_templates(request: Request, db: Session = Depends(get_db)):
    require_account_user(request, db)
    rows = db.query(AccountTemplate).filter(AccountTemplate.active.is_(True)).order_by(AccountTemplate.id.asc()).all()
    return [template_to_dict(row) for row in rows]


@router.post("/api/templates")
async def create_template(request: Request, db: Session = Depends(get_db)):
    require_account_user(request, db)
    data = await request.json()
    name = clean(data.get("template_name") or data.get("name"))
    if not name:
        raise HTTPException(status_code=400, detail="模板名称不能为空")
    if db.query(AccountTemplate).filter(AccountTemplate.template_name == name).first():
        raise HTTPException(status_code=400, detail="模板名称已存在")
    fields = data.get("fields") or ACCOUNT_HEADERS
    row = AccountTemplate(
        template_name=name,
        description=clean(data.get("description")),
        fields_json=json.dumps(filter_safe_export_headers(fields), ensure_ascii=False),
        active=True,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"ok": True, "id": row.id}


@router.get("/api/brands")
def list_brands(request: Request, db: Session = Depends(get_db)):
    require_account_user(request, db)
    rows = db.query(Brand).filter(Brand.active.is_(True)).order_by(Brand.brand_name.asc()).all()
    return [brand_to_dict(row) for row in rows]


@router.post("/api/brands")
async def create_brand(request: Request, db: Session = Depends(get_db)):
    require_account_user(request, db)
    data = await request.json()
    brand_name = clean(data.get("brand_name") or data.get("name"))
    if not brand_name:
        raise HTTPException(status_code=400, detail="品牌名称不能为空")
    brand = get_or_create_brand(db, brand_name)
    brand.brand_code = clean(data.get("brand_code")) or brand.brand_code
    brand.owner_name = clean(data.get("owner_name")) or brand.owner_name
    brand.notes = clean(data.get("notes")) or brand.notes
    db.commit()
    db.refresh(brand)
    return {"ok": True, "id": brand.id}


@router.get("/api/store-groups")
def list_store_groups(request: Request, db: Session = Depends(get_db)):
    require_account_user(request, db)
    rows = db.query(StoreGroup).filter(StoreGroup.active.is_(True)).order_by(StoreGroup.group_name.asc()).all()
    return [group_to_dict(row) for row in rows]


@router.post("/api/store-groups")
async def create_store_group(request: Request, db: Session = Depends(get_db)):
    require_account_user(request, db)
    data = await request.json()
    group_name = clean(data.get("group_name") or data.get("name"))
    if not group_name:
        raise HTTPException(status_code=400, detail="分组名称不能为空")
    group = get_or_create_group(db, group_name)
    group.group_code = clean(data.get("group_code")) or group.group_code
    group.description = clean(data.get("description")) or group.description
    db.commit()
    db.refresh(group)
    return {"ok": True, "id": group.id}


def upsert_account_meta(db: Session, store: Store, data: dict):
    brand_name = clean(data.get("brand") or data.get("品牌"))
    group_name = clean(data.get("group") or data.get("店铺分组"))
    owner = clean(data.get("owner") or data.get("负责人"))
    phone = clean(data.get("phone") or data.get("手机号"))
    status = clean(data.get("account_status") or data.get("status") or data.get("登录状态")) or "待登录"
    cookie_status = clean(data.get("cookie_status") or data.get("Cookie状态")) or status
    if status not in ACCOUNT_STATUSES:
        raise HTTPException(status_code=400, detail="登录状态必须是：正常、异常、待登录、Cookie过期、暂停采集")
    if cookie_status not in COOKIE_STATUSES:
        raise HTTPException(status_code=400, detail="Cookie状态不符合规范")

    note = db.query(StoreAccountNote).filter(StoreAccountNote.store_id == store.id).first()
    if not note:
        note = StoreAccountNote(store_id=store.id)
        db.add(note)
    note.brand = get_or_create_brand(db, brand_name) if brand_name else None
    note.group = get_or_create_group(db, group_name) if group_name else None
    note.login_account = clean(data.get("login_account") or data.get("登录账号"))
    plain_password = clean(data.get("plain_password"))
    if plain_password:
        note.encrypted_password = hash_password(plain_password)
    note.cookie_status = cookie_status
    note.login_status = status
    note.account_status = status
    note.last_login_at = parse_datetime(data.get("last_login_at") or data.get("最后登录时间"))
    note.notes = clean(data.get("notes") or data.get("备注"))
    note.tags = clean(data.get("tags") or data.get("标签"))
    note.active = True

    if owner:
        assignment = db.query(EmployeeStoreAssignment).filter(EmployeeStoreAssignment.store_id == store.id).first()
        if not assignment:
            assignment = EmployeeStoreAssignment(store_id=store.id, employee_name=owner)
            db.add(assignment)
        assignment.employee_name = owner
        assignment.phone = phone
        assignment.role_name = "负责人"
        assignment.active = True


def get_or_create_brand(db: Session, brand_name: str) -> Brand:
    row = db.query(Brand).filter(Brand.brand_name == brand_name).first()
    if row:
        return row
    row = Brand(brand_name=brand_name, active=True)
    db.add(row)
    db.flush()
    return row


def get_or_create_group(db: Session, group_name: str) -> StoreGroup:
    row = db.query(StoreGroup).filter(StoreGroup.group_name == group_name).first()
    if row:
        return row
    row = StoreGroup(group_name=group_name, active=True)
    db.add(row)
    db.flush()
    return row


def account_to_dict(note: StoreAccountNote, assignment: EmployeeStoreAssignment | None):
    return {
        "id": note.store_id,
        "store_id": note.store_id,
        "store_code": note.store.store_code if note.store else "",
        "store_name": note.store.store_name if note.store else "",
        "brand": note.brand.brand_name if note.brand else "",
        "brand_id": note.brand_id,
        "group": note.group.group_name if note.group else "",
        "group_id": note.group_id,
        "owner": assignment.employee_name if assignment else "",
        "phone": assignment.phone if assignment else "",
        "login_account": note.login_account or "",
        "cookie_status": note.cookie_status,
        "login_status": note.login_status,
        "account_status": note.account_status,
        "last_login_at": note.last_login_at.strftime("%Y-%m-%d %H:%M:%S") if note.last_login_at else "",
        "notes": note.notes or "",
        "tags": note.tags or "",
    }


def template_to_dict(row: AccountTemplate):
    return {
        "id": row.id,
        "template_name": row.template_name,
        "description": row.description or "",
        "fields": json.loads(row.fields_json or "[]"),
        "created_at": row.created_at.isoformat() if row.created_at else None,
    }


def brand_to_dict(row: Brand):
    return {
        "id": row.id,
        "brand_code": row.brand_code or "",
        "brand_name": row.brand_name,
        "owner_name": row.owner_name or "",
        "notes": row.notes or "",
    }


def group_to_dict(row: StoreGroup):
    return {
        "id": row.id,
        "group_code": row.group_code or "",
        "group_name": row.group_name,
        "description": row.description or "",
    }


def reject_forbidden_sensitive_payload(data: dict, allow_plain_password: bool = False):
    blocked = []
    for key in data:
        normalized = normalize_field_name(key)
        if allow_plain_password and normalized == "plain_password":
            continue
        if is_sensitive_field(normalized):
            blocked.append(str(key))
    if blocked:
        raise HTTPException(status_code=400, detail=f"敏感字段不允许直接提交：{', '.join(blocked)}")


def reject_sensitive_headers(headers: list[str]):
    blocked = [header for header in headers if is_sensitive_field(normalize_field_name(header))]
    if blocked:
        raise HTTPException(status_code=400, detail=f"Excel 不允许导入敏感字段：{', '.join(blocked)}")


def filter_safe_export_headers(headers: list[str]) -> list[str]:
    return [header for header in headers if not is_sensitive_field(normalize_field_name(header))]


def sanitize_export_row(row: dict) -> dict:
    return {key: clean(row.get(key, "")) for key in ALLOWED_EXPORT_KEYS}


def is_sensitive_field(normalized: str) -> bool:
    if normalized.endswith("_status") or normalized.endswith("status") or normalized.endswith("状态"):
        return False
    return any(keyword in normalized for keyword in SENSITIVE_EXCEL_HEADERS)


def normalize_field_name(value) -> str:
    return clean(value).lower().replace(" ", "").replace("-", "_")


def parse_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    text = clean(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise HTTPException(status_code=400, detail="最后登录时间格式应为 YYYY-MM-DD HH:mm:ss")


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def require_account_user(request: Request, db: Session):
    return require_permission_user(request, db, "stores.manage")
