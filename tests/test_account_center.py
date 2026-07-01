from io import BytesIO

from openpyxl import Workbook, load_workbook

from backend.models import Permission, Role, StoreAccountNote


def grant_account_center_access(session_factory):
    db = session_factory()
    try:
        permission = db.query(Permission).filter(Permission.code == "stores.manage").one_or_none()
        if not permission:
            permission = Permission(code="stores.manage", name="Manage stores")
            db.add(permission)
            db.flush()
        role = db.query(Role).filter(Role.code == "owner").one()
        if permission not in role.permissions:
            role.permissions.append(permission)
        db.commit()
    finally:
        db.close()


def login_owner(client):
    response = client.post("/api/login", json={"username": "owner", "password": "password"})
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['token']}"}


def make_xlsx(headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def assert_no_sensitive_payload(value):
    text = str(value).lower()
    for field in ("password", "encrypted_password", "token", "api_key", "secret"):
        assert field not in text
    assert "raw-cookie-value" not in text


def test_account_center_crud_import_export_and_sensitive_filtering(client, test_db):
    grant_account_center_access(test_db)
    headers = login_owner(client)

    blocked = client.post(
        "/api/accounts",
        headers=headers,
        json={
            "store_code": "JD999",
            "store_name": "Blocked Store",
            "encrypted_password": "do-not-save",
        },
    )
    assert blocked.status_code == 400

    created = client.post(
        "/api/accounts",
        headers=headers,
        json={
            "store_code": "JD101",
            "store_name": "JD Account Store",
            "brand": "Brand A",
            "owner": "Owner A",
            "phone": "13800000000",
            "login_account": "jd_user",
            "cookie_status": "正常",
            "account_status": "正常",
            "plain_password": "plain-secret",
            "notes": "status only",
            "tags": "key",
        },
    )
    assert created.status_code == 200
    account_id = created.json()["id"]

    db = test_db()
    try:
        note = db.query(StoreAccountNote).filter(StoreAccountNote.store_id == account_id).one()
        assert note.encrypted_password
        assert note.encrypted_password != "plain-secret"
        assert note.encrypted_password.startswith("pbkdf2_sha256$")
    finally:
        db.close()

    listed = client.get("/api/accounts", headers=headers)
    assert listed.status_code == 200
    rows = listed.json()
    assert rows[0]["store_code"] == "JD101"
    assert_no_sensitive_payload(rows)

    updated = client.put(
        f"/api/accounts/{account_id}",
        headers=headers,
        json={
            "store_code": "JD101",
            "store_name": "JD Account Store Updated",
            "brand": "Brand B",
            "owner": "Owner B",
            "cookie_status": "待登录",
            "account_status": "待登录",
        },
    )
    assert updated.status_code == 200

    bad_import_file = make_xlsx(
        ["店铺编号", "店铺名称", "password"],
        [["JD102", "Bad Import", "plain-secret"]],
    )
    bad_import = client.post(
        "/api/accounts/import",
        headers=headers,
        files={"file": ("bad.xlsx", bad_import_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert bad_import.status_code == 400

    import_file = make_xlsx(
        ["店铺编号", "店铺名称", "品牌", "负责人", "手机号", "登录账号", "Cookie状态", "登录状态", "备注", "标签"],
        [["JD102", "Imported Store", "Brand C", "Owner C", "13900000000", "import_user", "正常", "正常", "imported", "batch"]],
    )
    imported = client.post(
        "/api/accounts/import",
        headers=headers,
        files={"file": ("accounts.xlsx", import_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200
    assert imported.json()["imported"] == 1

    exported = client.get("/api/accounts/export", headers=headers)
    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content))
    sheet = workbook.active
    exported_values = [[cell.value for cell in row] for row in sheet.iter_rows()]
    assert_no_sensitive_payload(exported_values)
    assert "JD101" in str(exported_values)
    assert "JD102" in str(exported_values)

    deleted = client.delete(f"/api/accounts/{account_id}", headers=headers)
    assert deleted.status_code == 200


def test_low_privilege_user_cannot_access_account_center(client):
    viewer = client.post("/api/login", json={"username": "viewer", "password": "password"})
    assert viewer.status_code == 200
    headers = {"Authorization": f"Bearer {viewer.json()['token']}"}

    assert client.get("/api/accounts", headers=headers).status_code == 403
